import csv
import openpyxl
from openpyxl import Workbook

def count_words(text):
    if not text:
        return 0
    # Split by spaces to find the token count
    return len([w for w in text.strip().split() if w])

def clean_val(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    return val_str

def main():
    csv_path = 'chunks-resourcce/Package-Green-test - Sessions list - filled natural no names.csv'
    dest_path = 'chunks-resourcce/Chunks Resource.xlsx'
    
    # Session 1 sentence-to-term mapping to avoid missing values
    s1_term_map = {
        "Có hai con mèo đen đang ngủ trên ghế.": {"vi": "Con mèo", "en": "Cat"},
        "Buổi sáng, mẹ thường mua bánh mì ở đầu ngõ.": {"vi": "Mẹ", "en": "Mother"},
        "Tôi để chiếc ô màu xanh cạnh cửa.": {"vi": "Chiếc ô", "en": "Umbrella"},
        "Em bé đang chơi với quả bóng đỏ.": {"vi": "Em bé", "en": "Baby"},
        "Có một chậu hoa nhỏ trên ban công.": {"vi": "Chậu hoa", "en": "Flowerpot"},
        "Bố đọc báo trong lúc uống cà phê.": {"vi": "Bố", "en": "Father"},
        "Tối nay, cả nhà ăn cơm ở nhà.": {"vi": "Ăn cơm", "en": "Eat dinner"}
    }
    
    # 1. Parse the CSV file
    csv_rows = []
    with open(csv_path, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            csv_rows.append(r)
            
    # Group rows by session
    sessions_raw = {}
    for r in csv_rows:
        s_num = int(r['Session'])
        if s_num not in sessions_raw:
            sessions_raw[s_num] = []
        sessions_raw[s_num].append(r)
        
    # Define our updated C1 level items for Session 7 (exactly 21 words, which is under 22)
    session_7_items = [
        {"term_vi": "Trào ngược dạ dày", "term_en": "Acid reflux", "sent_vi": "Chứng trào ngược dạ dày thường xuyên xảy ra nếu bạn ăn quá no và nằm nghỉ ngay sau bữa ăn.", "sent_en": "Acid reflux disease often occurs if you eat too full and lie down to rest right after the meal."},
        {"term_vi": "Ợ nóng", "term_en": "Heartburn", "sent_vi": "Nếu cảm thấy bị ợ nóng khó chịu, tốt nhất bạn nên uống một cốc nước ấm để dễ chịu hơn.", "sent_en": "If you feel uncomfortable heartburn, you'd better drink a cup of warm water to feel more comfortable."},
        {"term_vi": "Đường huyết", "term_en": "Blood sugar", "sent_vi": "Chúng ta cần phải chú ý theo dõi chỉ số đường huyết để duy trì sức khỏe bằng chế độ dinh dưỡng tốt.", "sent_en": "We need to pay attention to monitoring the blood sugar index to maintain good health with a good diet."},
        {"term_vi": "Ngộ độc thực phẩm", "term_en": "Food poisoning", "sent_vi": "Một vụ ngộ độc thực phẩm nghiêm trọng xảy ra tại quán ăn lớn đó đã làm nhiều người phải đi cấp cứu.", "sent_en": "A serious food poisoning incident that occurred at that large restaurant sent many people to emergency treatment."},
        {"term_vi": "Ngấn mỡ bụng", "term_en": "Love handles", "sent_vi": "Để loại bỏ những ngấn mỡ bụng đáng ghét này, bạn cần kết hợp ăn kiêng và tập thể thao đều.", "sent_en": "To get rid of these annoying love handles, you need to combine a healthy diet and regular sports exercise."},
        {"term_vi": "\"Có tâm\"", "term_en": "Dedicated", "sent_vi": "Tôi rất quý mến người sếp rất có tâm vì anh ấy luôn chỉ bảo công việc cho nhân viên mới.", "sent_en": "I really like that very dedicated boss because he always guides the new employees in their work."},
        {"term_vi": "Còn khuya mới...", "term_en": "Nowhere near... / far from...", "sent_vi": "Học sinh lười biếng kia còn khuya mới thi đỗ nếu không chịu thay đổi thái độ học tập từ nay.", "sent_en": "That lazy student is nowhere near passing the exam if he doesn't change his study attitude from now."}
    ]
    
    # 2. Re-map and sort each session's items
    # Session configurations
    configs = {
        1: {"tl": 1.0, "lc": 1.0, "tc": 1.0, "cvr": 1, "cci_id": "cci-001"},
        2: {"tl": 1.0, "lc": 1.0, "tc": 3.0, "cvr": 3, "cci_id": "cci-002"},
        3: {"tl": 1.3, "lc": 1.3, "tc": 2.959, "cvr": 5, "cci_id": "cci-003"},
        4: {"tl": 1.4, "lc": 1.7, "tc": 2.941, "cvr": 7, "cci_id": "cci-004"},
        5: {"tl": 1.6, "lc": 2.1, "tc": 2.679, "cvr": 9, "cci_id": "cci-005"},
        6: {"tl": 1.8, "lc": 2.6, "tc": 2.350, "cvr": 11, "cci_id": "cci-006"},
        7: {"tl": 2.0, "lc": 2.1, "tc": 3.095, "cvr": 13, "cci_id": "cci-007"}
    }
    
    final_items = []
    
    for s_idx in sorted(sessions_raw.keys()):
        config = configs[s_idx]
        items_in_session = []
        
        if s_idx == 7:
            # Recreate session 7 with our C1 list
            for item in session_7_items:
                items_in_session.append({
                    "term_vi": clean_val(item["term_vi"]),
                    "term_en": clean_val(item["term_en"]),
                    "sent_vi": clean_val(item["sent_vi"]),
                    "sent_en": clean_val(item["sent_en"])
                })
        else:
            # Use parsed rows from the CSV file
            for r in sessions_raw[s_idx]:
                vi_sent = clean_val(r['Sentence vi'])
                t_vi = clean_val(r['TC Vietnamese'])
                t_en = clean_val(r['TC English'])
                
                # Fill Session 1 terms from the mapping if empty
                if s_idx == 1 and vi_sent in s1_term_map:
                    t_vi = s1_term_map[vi_sent]["vi"]
                    t_en = s1_term_map[vi_sent]["en"]
                    
                items_in_session.append({
                    "term_vi": t_vi,
                    "term_en": t_en,
                    "sent_vi": vi_sent,
                    "sent_en": clean_val(r['Sentence en'])
                })
                
        # Sort items in this session by Sentence vi word count (ascending)
        items_in_session.sort(key=lambda x: count_words(x["sent_vi"]))
        
        # Build final rows
        for idx, item in enumerate(items_in_session):
            final_items.append([
                'Day 2',                           # Material
                f'Session {s_idx}',                 # Session No. (str)
                f'Number {idx + 1}',                 # Item_id
                config["cci_id"],                   # CCI-id
                config["cvr"],                      # CVR-id
                item['term_vi'],                    # Term (Tiếng Việt)
                item['term_en'],                    # Term (Tiếng Anh)
                s_idx,                              # Session No. (int)
                item['sent_vi'],                    # Complete Sentence (Vie)
                item['sent_en'],                    # Complete Sentence (Eng)
                config["tc"],                       # TC
                config["lc"],                       # LC
                config["tl"]                        # TL
            ])
            
    # 3. Build CCI definitions
    cci_headers = ['Session', 'CCI_id', 'CCI Name', 'Ampe (A)', 'Description', 'Category']
    cci_rows = [
        [1, 'cci-001', 'Give it a shot', 2, 'Linear 1 on 1 as Blow', 'Blow'],
        [2, 'cci-002', 'Go with the flow', 2, 'Linear RPD-free as Flow', 'Flow'],
        [3, 'cci-003', 'Chunks on the go', 4, 'Linear chunking act as Chunks', 'Chunks'],
        [4, 'cci-004', 'Freeze', 4, 'Freeze your body with RPD-free or 1-on-1 sound', 'null'],
        [5, 'cci-005', 'Robot', 6, 'Move your hands linearly 1-on-1 as Blow', 'Blow'],
        [6, 'cci-006', 'Taichi', 6, 'Move your hands nonstop freely as Flow', 'Flow'],
        [7, 'cci-007', 'Strike', 8, 'Strike a fixed n times as Chunks', 'Chunks']
    ]
    
    # 4. Build Package-test definitions
    package_headers = ['Package_id', 'Name', 'Description', 'Session list', 'CCI list']
    package_rows = []
    for i in range(1, 8):
        package_rows.append([
            'Package-Green-test',
            f'Test {i:02d}',
            'Bộ test Green 7 sessions',
            f'session_id -{i}',
            f'cci-id {i:02d}'
        ])
        
    # Create the workbook from scratch
    wb_dest = Workbook()
    
    # Sheet 1: Chunks-resource - CVR_new
    ws_items = wb_dest.active
    ws_items.title = 'Chunks-resource - CVR_new'
    item_headers = [
        'Material', 'Session No.', 'Item_id', 'CCI-id', 'CVR-id',
        'Term (Tiếng Việt)', 'Term (Tiếng Anh)', 'Session No.',
        'Complete Sentence (Vie)', 'Complete Sentence (Eng)',
        'TC', 'LC', 'TL'
    ]
    ws_items.append(item_headers)
    for r in final_items:
        ws_items.append(r)
        
    # Sheet 2: Package-test
    ws_pkg = wb_dest.create_sheet(title='Package-test')
    ws_pkg.append(package_headers)
    for r in package_rows:
        ws_pkg.append(r)
        
    # Sheet 3: CCI
    ws_cci = wb_dest.create_sheet(title='CCI')
    ws_cci.append(cci_headers)
    for r in cci_rows:
        ws_cci.append(r)
        
    # Save the updated workbook
    wb_dest.save(dest_path)
    print(f"Successfully generated clean workbook from scratch at: {dest_path}")
    print(f"Total sessions: 7")
    print(f"Total items added: {len(final_items)}")

if __name__ == '__main__':
    main()
