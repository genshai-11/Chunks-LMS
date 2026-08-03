import csv
import openpyxl
from openpyxl import Workbook

def count_words(text):
    if not text:
        return 0
    # Split by spaces to find the token count
    return len([w for w in text.strip().split() if w])

def clean_val(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    return val_str

def main():
    csv_path = 'chunks-resourcce/Package-Green-test - Sessions list - option 2.csv'
    dest_path = 'chunks-resourcce/Chunks Resource Option 2.xlsx'
    
    # 1. Option 2 terms and sentences with increased TL (Vocabulary Level)
    # Session 1: Advanced terms (Phát triển, Công nghệ, Mục tiêu, Phương pháp, Trách nhiệm, Quyết định, Kiến thức)
    # Exactly 10 words per sentence
    session_1_items = [
        {"item": 1, "term_vi": "Phát triển", "term_en": "Development", "sent_vi": "Công ty luôn tập trung phát triển nguồn nhân lực trẻ.", "sent_en": "The company always focuses on developing young human resources."},
        {"item": 2, "term_vi": "Công nghệ", "term_en": "Technology", "sent_vi": "Công nghệ hiện đại giúp nâng cao hiệu quả làm việc.", "sent_en": "Modern technology helps to improve this working efficiency."},
        {"item": 3, "term_vi": "Mục tiêu", "term_en": "Goal / Target", "sent_vi": "Mỗi người nên tự đặt ra mục tiêu của mình.", "sent_en": "Each person should set goals for themselves."},
        {"item": 4, "term_vi": "Phương pháp", "term_en": "Method", "sent_vi": "Phương pháp giảng dạy mới mang lại hiệu quả cao.", "sent_en": "The new teaching method brings about high efficiency."},
        {"item": 5, "term_vi": "Trách nhiệm", "term_en": "Responsibility", "sent_vi": "Mỗi nhân viên cần có trách nhiệm với công việc.", "sent_en": "Each employee needs to be responsible for their work."},
        {"item": 6, "term_vi": "Quyết định", "term_en": "Decision", "sent_vi": "Ban giám đốc vừa đưa ra quyết định cuối cùng.", "sent_en": "The board of directors has just made the final decision."},
        {"item": 7, "term_vi": "Kiến thức", "term_en": "Knowledge", "sent_vi": "Đọc sách giúp chúng ta tích lũy nhiều kiến thức.", "sent_en": "Reading books helps us accumulate a lot of knowledge."}
    ]
    
    # Session 2: Same terms, different sentences (9-10 words)
    session_2_items = [
        {"item": 1, "term_vi": "Duyệt", "term_en": "Approve / give me a go-ahead", "sent_vi": "Giám đốc duyệt phương án quảng cáo mới rồi.", "sent_en": "The director approved the new advertising plan."},
        {"item": 2, "term_vi": "..., có thể nói vậy", "term_en": "..., so to speak", "sent_vi": "Kế hoạch đã thất bại, có thể nói vậy.", "sent_en": "The plan has failed, so to speak."},
        {"item": 3, "term_vi": "Cải thiện", "term_en": "Work on / improve", "sent_vi": "Chúng tôi muốn cải thiện chất lượng dịch vụ.", "sent_en": "We want to improve our service quality."},
        {"item": 4, "term_vi": "Quản trị", "term_en": "Management", "sent_vi": "Anh ấy đang học quản trị nhân sự mới.", "sent_en": "He is studying new human resource management."},
        {"item": 5, "term_vi": "Nguyên tắc", "term_en": "Philosophy / principle", "sent_vi": "Làm việc nhóm luôn cần có nguyên tắc chung.", "sent_en": "Teamwork always needs to have common principles."},
        {"item": 6, "term_vi": "Dài dòng", "term_en": "Lengthy / wordy", "sent_vi": "Bài thuyết trình dài dòng gây buồn ngủ quá.", "sent_en": "The lengthy presentation makes me so sleepy."},
        {"item": 7, "term_vi": "Dầu gió", "term_en": "Medicated oil", "sent_vi": "Mẹ hay dùng dầu gió khi bị cảm lạnh.", "sent_en": "Mom often uses medicated oil when she catches a cold."}
    ]
    
    # Session 3: Same terms, different sentences (12 words)
    session_3_items = [
        {"item": 1, "term_vi": "Cái máy lạnh", "term_en": "AC / Air Conditioner", "sent_vi": "Tôi mở cái máy lạnh vì thời tiết hôm nay quá oi bức.", "sent_en": "I turned on the air conditioner because the weather is too hot today."},
        {"item": 2, "term_vi": "Cứ thoải mái", "term_en": "Freely / feel free", "sent_vi": "Bạn cứ thoải mái chọn món ăn ưa thích trong thực đơn này.", "sent_en": "Please feel free to choose your favorite dish from this menu."},
        {"item": 3, "term_vi": "Kỹ sư trưởng", "term_en": "Chief engineer", "sent_vi": "Chú tôi làm kỹ sư trưởng ở công trường xây dựng này.", "sent_en": "My uncle works as a chief engineer at this construction site."},
        {"item": 4, "term_vi": "Kỹ năng và khả năng", "term_en": "Skill and ability", "sent_vi": "Khóa học này rèn luyện kỹ năng và khả năng tự học.", "sent_en": "This course trains self-study skills and abilities."},
        {"item": 5, "term_vi": "Thực tập sinh", "term_en": "Intern / trainee / probationer / apprentice", "sent_vi": "Tôi hướng dẫn thực tập sinh mới dùng máy in văn phòng.", "sent_en": "I guide the new intern to use the office printer."},
        {"item": 6, "term_vi": "Có lẽ", "term_en": "Perhaps / maybe / most likely", "sent_vi": "Có lẽ chúng ta nên dời cuộc họp sang sáng mai nhé.", "sent_en": "Perhaps we should move the meeting to tomorrow morning."},
        {"item": 7, "term_vi": "Duyệt", "term_en": "Approve / give me a go-ahead", "sent_vi": "Giám đốc đã duyệt đơn xin nghỉ phép của tôi chiều nay.", "sent_en": "The director approved my leave request this afternoon."}
    ]
    
    # Session 4: Same terms, different sentences (13 words)
    session_4_items = [
        {"item": 1, "term_vi": "Dễ (kiếm tiền)", "term_en": "Easy money", "sent_vi": "Làm khảo sát trực tuyến không phải cách dễ kiếm tiền đâu bạn.", "sent_en": "Doing online surveys is not an easy way to make money, my friend."},
        {"item": 2, "term_vi": "Hợp đồng", "term_en": "Contract", "sent_vi": "Luật sư đang kiểm tra lại các điều khoản của bản hợp đồng.", "sent_en": "The lawyer is checking the terms of the contract again."},
        {"item": 3, "term_vi": "Ký một cái hợp đồng", "term_en": "Sign a contract", "sent_vi": "Chúng tôi chuẩn bị ký một cái hợp đồng thuê nhà dài hạn.", "sent_en": "We are preparing to sign a long-term house rental contract."},
        {"item": 4, "term_vi": "Trước tiên", "term_en": "First / firstly / first off", "sent_vi": "Trước tiên hãy rửa tay sạch sẽ trước khi bắt đầu bữa ăn.", "sent_en": "First of all, wash your hands clean before starting the meal."},
        {"item": 5, "term_vi": "Thương lượng", "term_en": "Negotiate", "sent_vi": "Người mua đang thương lượng với chủ nhà để giảm giá thuê phòng.", "sent_en": "The buyer is negotiating with the landlord to reduce room rent."},
        {"item": 6, "term_vi": "Ồn ào / to mồm", "term_en": "Noisy / loud-mouthed", "sent_vi": "Tiếng xe cộ ồn ào ngoài đường làm em bé thức giấc rồi.", "sent_en": "The noisy traffic noise outside woke the baby up already."},
        {"item": 7, "term_vi": "Đồng nghiệp", "term_en": "Coworker / colleague / office buddy", "sent_vi": "Tôi vừa cùng đồng nghiệp ăn trưa tại quán ăn đối diện này.", "sent_en": "I just had lunch with my colleagues at the opposite restaurant."}
    ]
    
    # Session 5: Same terms, different sentences (17 words)
    session_5_items = [
        {"item": 1, "term_vi": "Nhiều khi / có mấy lúc", "term_en": "Sometimes / once in a while / at times", "sent_vi": "Nhiều khi tôi thấy nhớ quê hương và muốn bắt chuyến xe về thăm gia đình mình.", "sent_en": "Sometimes I miss my hometown and want to catch a bus to visit my family."},
        {"item": 2, "term_vi": "Tôi e là không", "term_en": "I'm afraid not", "sent_vi": "Tôi e là không thể hoàn thành việc sửa xe này trước khi trời tối đâu nhé.", "sent_en": "I'm afraid I cannot finish repairing this car before dark."},
        {"item": 3, "term_vi": "(Ý anh) là sao?", "term_en": "What'd you mean?", "sent_vi": "Ý anh là sao khi nói rằng chúng ta nên hoãn chuyến du lịch cuối tuần này?", "sent_en": "What do you mean by saying we should postpone our trip this weekend?"},
        {"item": 4, "term_vi": "Phức tạp", "term_en": "Complicated / complex / like a mess", "sent_vi": "Cách sử dụng phần mềm mới này khá phức tạp đối với những người lớn tuổi rồi.", "sent_en": "How to use this new software is quite complicated for elderly people."},
        {"item": 5, "term_vi": "Mỗi 5 năm", "term_en": "Every 5 years", "sent_vi": "Mỗi năm năm một lần, trường cũ của tôi lại tổ chức ngày hội ngộ cựu học sinh.", "sent_en": "Once every five years, my old school organizes an alumni reunion day."},
        {"item": 6, "term_vi": "Thông thường", "term_en": "Usually / often / always", "sent_vi": "Thông thường tôi sẽ đi ngủ sớm để giữ sức khỏe tốt cho ngày làm việc sau.", "sent_en": "Usually I will go to bed early to keep good health for the next working day."},
        {"item": 7, "term_vi": "Vấn đề thiếu hụt nhân lực", "term_en": "Staff shortage / headcount problem", "sent_vi": "Vấn đề thiếu hụt nhân lực làm dự án bị chậm tiến độ so với kế hoạch.", "sent_en": "The staff shortage problem makes the project delayed compared to the plan."}
    ]
    
    # Session 6: Same terms, different sentences (15 words)
    session_6_items = [
        {"item": 1, "term_vi": "Tốt hơn là cậu nên…", "term_en": "You'd better… / You should...", "sent_vi": "Tốt hơn là cậu nên tắt máy tính và đi ngủ sớm đi nghe chưa.", "sent_en": "You had better turn off the computer and go to bed early."},
        {"item": 2, "term_vi": "Chuỗi cung ứng", "term_en": "Supply chain", "sent_vi": "Sự tắc nghẽn giao thông làm ảnh hưởng chuỗi cung ứng hàng hóa này rồi.", "sent_en": "The traffic congestion has affected this goods supply chain already."},
        {"item": 3, "term_vi": "Thiếu hụt", "term_en": "Shortage / scarcity", "sent_vi": "Tình trạng thiếu hụt năng lượng đang xảy ra ở nhiều nước châu Âu này.", "sent_en": "The energy shortage situation is happening in many European countries."},
        {"item": 4, "term_vi": "Đại dịch", "term_en": "Pandemic / epidemic", "sent_vi": "Mọi người phải đeo khẩu trang khi đi xe buýt trong thời kỳ đại dịch.", "sent_en": "People had to wear masks when riding the bus during the pandemic."},
        {"item": 5, "term_vi": "Chính phủ Trung Quốc", "term_en": "Chinese government", "sent_vi": "Chính phủ Trung Quốc đã đầu tư nhiều tiền vào nghiên cứu vũ trụ mới.", "sent_en": "The Chinese government has invested a lot of money in new space research."},
        {"item": 6, "term_vi": "đóng cửa", "term_en": "Shut down", "sent_vi": "Khu vui chơi trẻ em đóng cửa để sửa chữa các thiết bị hỏng này.", "sent_en": "The children's playground is closed to repair these broken devices."},
        {"item": 7, "term_vi": "Nỗi đau thật sự", "term_en": "Real pain point", "sent_vi": "Việc mất dữ liệu khách hàng chính là nỗi đau thật sự của sếp tôi.", "sent_en": "Losing customer database is the real pain point of my boss."}
    ]
    
    # Session 7: Same terms, different sentences (21 words, C1 level)
    session_7_items = [
        {"item": 1, "term_vi": "Trào ngược dạ dày", "term_en": "Acid reflux", "sent_vi": "Việc lạm dụng các chất kích thích có thể làm trầm trọng thêm tình trạng trào ngược dạ dày bệnh nhân.", "sent_en": "Abusing stimulants can worsen the acid reflux condition of the patient."},
        {"item": 2, "term_vi": "Ợ nóng", "term_en": "Heartburn", "sent_vi": "Chế độ dinh dưỡng không lành mạnh là một trong những nguyên nhân hàng đầu gây ra chứng bệnh lý này.", "sent_en": "An unhealthy diet is one of the leading causes of this pathological condition."},
        {"term_vi": "Đường huyết", "term_en": "Blood sugar", "sent_vi": "Các bác sĩ khuyến cáo bệnh nhân cần duy trì thói quen đo chỉ số đường huyết định kỳ hằng ngày.", "sent_en": "Doctors recommend that patients need to maintain a habit of measuring blood sugar levels daily."},
        {"item": 4, "term_vi": "Ngộ độc thực phẩm", "term_en": "Food poisoning", "sent_vi": "Cơ quan y tế đang nỗ lực điều tra nguyên nhân gây ra vụ ngộ độc thực phẩm lớn vừa qua.", "sent_en": "The health agency is striving to investigate the cause of the recent large-scale food poisoning incident."},
        {"item": 5, "term_vi": "Ngấn mỡ bụng", "term_en": "Love handles", "sent_vi": "Tình trạng tích tụ mỡ thừa tạo thành các ngấn mỡ bụng thường xuất phát từ lối sống ít vận động.", "sent_en": "The accumulation of excess fat forming love handles often stems from a sedentary lifestyle."},
        {"item": 6, "term_vi": "\"Có tâm\"", "term_en": "Dedicated", "sent_vi": "Một người thầy thuốc có tâm luôn đặt sức khỏe và lợi ích người bệnh lên trên hết thảy việc khác.", "sent_en": "A dedicated physician always puts the health and interests of patients above all else."},
        {"item": 7, "term_vi": "Còn khuya mới...", "term_en": "Nowhere near... / far from...", "sent_vi": "Tổ chức của chúng ta còn khuya mới vượt qua khủng hoảng nếu ban lãnh đạo không cải tiến tư duy.", "sent_en": "Our organization is nowhere near overcoming the crisis if the leadership does not improve thinking."}
    ]
    
    # 2. Write this Option 2 data directly into the CSV file
    csv_rows = [
        ["", "Session ", "Item", "CCI", "CVR", "TC Vietnamese", "TC English", "Sentence vi", "Sentence en"]
    ]
    
    # Session mapping helper
    sessions_map = {
        1: session_1_items,
        2: session_2_items,
        3: session_3_items,
        4: session_4_items,
        5: session_5_items,
        6: session_6_items,
        7: session_7_items
    }
    
    ccis = {1: 2, 2: 2, 3: 4, 4: 4, 5: 6, 6: 6, 7: 8}
    cvrs = {1: 1, 2: 3, 3: 5, 4: 7, 5: 9, 6: 11, 7: 13}
    
    for s_idx in range(1, 8):
        items = sessions_map[s_idx]
        cci = ccis[s_idx]
        cvr = cvrs[s_idx]
        for item in items:
            csv_rows.append([
                "",
                str(s_idx),
                str(item["item"]),
                str(cci),
                str(cvr),
                item["term_vi"],
                item["term_en"],
                item["sent_vi"],
                item["sent_en"]
            ])
            
    # Write CSV file
    # Prepend empty row at index 0 to match exactly the 9-column CSV format of the original file
    final_csv_rows = [["", "", "", "", "", "", "", "", ""]] + csv_rows
    
    with open(csv_path, mode='w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(final_csv_rows)
        
    print(f"Successfully created CSV Option 2 at: {csv_path}")
    
    # 3. Now read the CSV, sort items by word count and generate Chunks Resource Option 2.xlsx
    configs = {
        1: {"tl": 1.0, "lc": 1.0, "tc": 1.0, "cvr": 1, "cci_id": "cci-001"},
        2: {"tl": 1.0, "lc": 1.0, "tc": 3.0, "cvr": 3, "cci_id": "cci-002"},
        3: {"tl": 1.3, "lc": 1.3, "tc": 2.959, "cvr": 5, "cci_id": "cci-003"},
        4: {"tl": 1.4, "lc": 1.7, "tc": 2.941, "cvr": 7, "cci_id": "cci-004"},
        5: {"tl": 1.6, "lc": 2.1, "tc": 2.679, "cvr": 9, "cci_id": "cci-005"},
        6: {"tl": 1.8, "lc": 2.6, "tc": 2.350, "cvr": 11, "cci_id": "cci-006"},
        7: {"tl": 2.0, "lc": 2.1, "tc": 3.095, "cvr": 13, "cci_id": "cci-007"}
    }
    
    final_items = []
    
    # Read back the CSV to perform length sorting
    sessions_raw = {}
    for row in csv_rows[1:]: # Skip headers
        s_num = int(row[1])
        if s_num not in sessions_raw:
            sessions_raw[s_num] = []
        sessions_raw[s_num].append({
            "term_vi": row[5],
            "term_en": row[6],
            "sent_vi": row[7],
            "sent_en": row[8]
        })
        
    for s_idx in sorted(sessions_raw.keys()):
        config = configs[s_idx]
        items_in_session = sessions_raw[s_idx]
        
        # Sort items in this session by Sentence vi word count (ascending)
        items_in_session.sort(key=lambda x: count_words(x["sent_vi"]))
        
        # Build final rows
        for idx, item in enumerate(items_in_session):
            final_items.append([
                'Day 2',                           # Material
                f'Session {s_idx}',                 # Session No. (str)
                f'Number {idx + 1}',                 # Item_id
                config["cci_id"],                   # CCI-id
                config["cvr"],                      # CVR-id
                clean_val(item['term_vi']),         # Term (Tiếng Việt)
                clean_val(item['term_en']),         # Term (Tiếng Anh)
                s_idx,                              # Session No. (int)
                clean_val(item['sent_vi']),         # Complete Sentence (Vie)
                clean_val(item['sent_en']),         # Complete Sentence (Eng)
                config["tc"],                       # TC
                config["lc"],                       # LC
                config["tl"]                        # TL
            ])
            
    # CCI definitions
    cci_headers = ['Session', 'CCI_id', 'CCI Name', 'Ampe (A)', 'Description', 'Category']
    cci_rows = [
        [1, 'cci-001', 'Give it a shot', 2, 'Linear 1 on 1 as Blow', 'Blow'],
        [2, 'cci-002', 'Go with the flow', 2, 'Linear RPD-free as Flow', 'Flow'],
        [3, 'cci-003', 'Chunks on the go', 4, 'Linear chunking act as Chunks', 'Chunks'],
        [4, 'cci-004', 'Freeze', 4, 'Freeze your body with RPD-free or 1-on-1 sound', 'null'],
        [5, 'cci-005', 'Robot', 6, 'Move your hands linearly 1-on-1 as Blow', 'Blow'],
        [6, 'cci-006', 'Taichi', 6, 'Move your hands nonstop freely as Flow', 'Flow'],
        [7, 'cci-007', 'Strike', 8, 'Strike a fixed n times as Chunks', 'Chunks']
    ]
    
    # Package-test definitions
    package_headers = ['Package_id', 'Name', 'Description', 'Session list', 'CCI list']
    package_rows = []
    for i in range(1, 8):
        package_rows.append([
            'Package-Green-test-opt2',
            f'Test {i:02d}',
            'Bộ test Green Option 2',
            f'session_id -{i}',
            f'cci-id {i:02d}'
        ])
        
    # Create the workbook from scratch
    wb_dest = Workbook()
    
    # Sheet 1: Chunks-resource - CVR_new
    ws_items = wb_dest.active
    ws_items.title = 'Chunks-resource - CVR_new'
    item_headers = [
        'Material', 'Session No.', 'Item_id', 'CCI-id', 'CVR-id',
        'Term (Tiếng Việt)', 'Term (Tiếng Anh)', 'Session No.',
        'Complete Sentence (Vie)', 'Complete Sentence (Eng)',
        'TC', 'LC', 'TL'
    ]
    ws_items.append(item_headers)
    for r in final_items:
        ws_items.append(r)
        
    # Sheet 2: Package-test
    ws_pkg = wb_dest.create_sheet(title='Package-test')
    ws_pkg.append(package_headers)
    for r in package_rows:
        ws_pkg.append(r)
        
    # Sheet 3: CCI
    ws_cci = wb_dest.create_sheet(title='CCI')
    ws_cci.append(cci_headers)
    for r in cci_rows:
        ws_cci.append(r)
        
    wb_dest.save(dest_path)
    print(f"Successfully generated Option 2 workbook at: {dest_path}")

if __name__ == '__main__':
    main()
