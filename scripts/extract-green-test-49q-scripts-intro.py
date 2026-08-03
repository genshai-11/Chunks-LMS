from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook

root = Path.cwd()
workbook = root / 'chunks-resourcce' / 'Package-Green-test.xlsx'
out = root / 'supabase' / 'seeds' / 'green-test-49q' / 'GREEN-TEST-49Q.scripts-intro.json'
wb = load_workbook(workbook, read_only=True, data_only=True)
ws = wb['Scripts Intro']
rows = []
for row_index, row in enumerate(ws.iter_rows(values_only=True), 1):
    key = row[0] if len(row) > 0 else None
    text = row[1] if len(row) > 1 else None
    if key and text:
        rows.append({'row': row_index, 'key': str(key).strip(), 'text': str(text).strip()})
expected = [
    'audio_intro_test', 'audio_intro_part_I', 'audio_intro_part_II', 'audio_intro_part_III', 'audio_end_test',
    'audio_intro_session1', 'audio_intro_session2', 'audio_intro_session3', 'audio_intro_session4',
    'audio_intro_session5', 'audio_intro_session6', 'audio_intro_session7',
]
keys = [r['key'] for r in rows]
missing = [k for k in expected if k not in keys]
extra = [k for k in keys if k not in expected]
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps({'sourceWorkbook': str(workbook), 'sheet': 'Scripts Intro', 'count': len(rows), 'missing': missing, 'extra': extra, 'rows': rows}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({'out': str(out), 'count': len(rows), 'missing': missing, 'extra': extra}, ensure_ascii=False, indent=2))
if missing:
    raise SystemExit(1)
