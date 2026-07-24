from __future__ import annotations
import hashlib
import json
from pathlib import Path
from openpyxl import load_workbook

root = Path.cwd()
workbook = root / 'chunks-resourcce' / 'Package-Green-test.xlsx'
out = root / 'supabase' / 'seeds' / 'green-test-49q' / 'GREEN-TEST-49Q.workbook-manifest.json'
wb = load_workbook(workbook, read_only=True, data_only=True)

# Scripts Intro
ws_intro = wb['Scripts Intro']
scripts_intro = {}
script_rows = []
for row_index, row in enumerate(ws_intro.iter_rows(values_only=True), 1):
    key = row[0] if len(row) > 0 else None
    text = row[1] if len(row) > 1 else None
    if key and text:
        key = str(key).strip()
        text = str(text).strip()
        scripts_intro[key] = text
        script_rows.append({'row': row_index, 'key': key, 'text': text})

# CCI definitions from right-side table in Script session intro
ws_cci = wb['Script session intro']
cci_definitions = []
for row in ws_cci.iter_rows(min_row=4, max_row=10, values_only=True):
    session = row[3]
    if not session:
        continue
    cci_definitions.append({
        'sessionOrder': int(session),
        'sourceCciId': f"green49q-cci-{int(session):03d}",
        'name': str(row[4]).strip(),
        'targetCvrOhm': float(row[5]),
        'ampe': float(row[6]),
        'cpd': float(row[7]),
        'description': str(row[8]).strip() if row[8] is not None else None,
        'category': None,
    })

# Package per-session metrics
ws_metrics = wb['Package CVR CCI CPD']
metrics = {}
for row in ws_metrics.iter_rows(min_row=3, max_row=9, values_only=True):
    if row[1] is None:
        continue
    session = int(row[1])
    metrics[session] = {'cvr': float(row[2]), 'cci': float(row[3]), 'cpd': float(row[4])}

# Sessions list items
ws = wb['Sessions list']
sessions = {i: [] for i in range(1, 8)}
for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
    session = row[1] if len(row) > 1 else None
    item = row[2] if len(row) > 2 else None
    if session is None or item is None:
        continue
    session = int(session)
    item_order = int(item)
    cci = float(row[3]) if row[3] is not None else None
    cvr = float(row[4]) if row[4] is not None else None
    term_vi = str(row[5]).strip() if row[5] is not None else None
    term_en = str(row[6]).strip() if row[6] is not None else None
    prompt_vi = str(row[7]).strip() if row[7] is not None else None
    prompt_en = str(row[8]).strip() if row[8] is not None else None
    sessions[session].append({
        'itemOrder': item_order,
        'sourceItemId': f'Number {item_order}',
        'sourceMaterial': 'Package-Green-test.xlsx',
        'sourceCciId': f"green49q-cci-{session:03d}",
        'sourceCvrId': cvr,
        'termVi': term_vi,
        'termEn': term_en,
        'promptVi': prompt_vi,
        'promptEn': prompt_en,
        'tc': cvr,
        'lc': 1,
        'tl': 1,
        'sourceRow': row_index,
        'wordCountVi': row[9] if len(row) > 9 else None,
        'wordCountEn': row[10] if len(row) > 10 else None,
    })

session_objs = []
for session_order in range(1, 8):
    cci = next(c for c in cci_definitions if c['sessionOrder'] == session_order)
    intro = scripts_intro[f'audio_intro_session{session_order}']
    session_objs.append({
        'sessionOrder': session_order,
        'name': f'Test {session_order:02d}',
        'description': 'GREEN-TEST-49Q from Package-Green-test.xlsx',
        'sourceCciId': cci['sourceCciId'],
        'targetCvrOhm': metrics.get(session_order, {}).get('cvr', cci['targetCvrOhm']),
        'introTextVi': intro,
        'introTextEn': intro,
        'items': sorted(sessions[session_order], key=lambda x: x['itemOrder']),
    })

manifest = {
    'source': {
        'filename': workbook.name,
        'sha256': hashlib.sha256(workbook.read_bytes()).hexdigest(),
        'sheets': ['Sessions list', 'Package CVR CCI CPD', 'Script session intro', 'Scripts Intro'],
    },
    'package': {
        'sourcePackageId': 'GREEN-TEST-49Q',
        'title': 'GREEN-TEST-49Q',
        'description': 'GREEN-TEST-49Q · Package-Green-test.xlsx · 7 sessions × 7 questions',
        'versionLabel': 'LIVE',
    },
    'scriptsIntro': script_rows,
    'cciDefinitions': cci_definitions,
    'sessions': session_objs,
    'issues': [],
}

out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'out': str(out),
    'sourceSha256': manifest['source']['sha256'],
    'sessions': len(session_objs),
    'items': sum(len(s['items']) for s in session_objs),
    'scriptsIntro': len(script_rows),
    'cciDefinitions': len(cci_definitions),
}, ensure_ascii=False, indent=2))
if len(session_objs) != 7 or sum(len(s['items']) for s in session_objs) != 49 or len(script_rows) != 12:
    raise SystemExit(1)
