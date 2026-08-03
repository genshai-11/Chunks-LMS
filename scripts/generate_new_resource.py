import openpyxl
from openpyxl import Workbook
import os

def main():
    # Paths
    src_path = 'chunks-resourcce/Chunks Resource.xlsx'
    dest_path = 'chunks-resourcce/Chunks Resource New.xlsx'
    
    if not os.path.exists(src_path):
        print(f"Error: Source workbook {src_path} not found.")
        return
        
    wb_src = openpyxl.load_workbook(src_path)
    
    # 1. Prepare new CCI Sheet Data
    # Columns: ['Session', 'CCI_id', 'CCI Name', 'Ampe (A)', 'Description', 'Category']
    cci_headers = ['Session', 'CCI_id', 'CCI Name', 'Ampe (A)', 'Description', 'Category']
    cci_rows = [
        [1, 'cci-001', 'Give it a shot', 2, 'Linear 1 on 1 as Blow', 'Blow'],
        [2, 'cci-002', 'Go with the flow', 2, 'Linear RPD-free as Flow', 'Flow'],
        [3, 'cci-003', 'Chunks on the go', 4, 'Linear chunking act as Chunks', 'Chunks'],
        [4, 'cci-004', 'Robot', 6, 'Move your hands linearly 1-on-1 as Blow', 'Blow'],
        [5, 'cci-005', 'Taichi', 6, 'Move your hands nonstop freely as Flow', 'Flow'],
        [6, 'cci-006', 'Strike', 8, 'Strike a fixed n times as Chunks', 'Chunks'],
        [7, 'cci-007', 'Combine', 9, 'Combine (Freeze + Nuance Work)', 'null']
    ]
    
    # 2. Prepare Package-test Sheet Data
    # Columns: ['Package_id', 'Name', 'Description', 'Session list', 'CCI list']
    package_headers = ['Package_id', 'Name', 'Description', 'Session list', 'CCI list']
    package_rows = []
    for i in range(1, 8):
        package_rows.append([
            'New-test-package',
            f'Test {i:02d}',
            'Bộ test mới 7 sessions',
            f'session_id -{i}',
            f'cci-id {i:02d}'
        ])
        
    # 3. Prepare Chunks-resource - CVR_new Sheet Data
    item_headers = [
        'Material',
        'Session No.',
        'Item_id',
        'CCI-id',
        'CVR-id',
        'Term (Tiếng Việt)',
        'Term (Tiếng Anh)',
        'Session No.',
        'Complete Sentence (Vie)',
        'Complete Sentence (Eng)',
        'TC',
        'LC',
        'TL'
    ]
    
    # Define exact sentences and parameters according to the redefined rules:
    # Session 1: LC = 1 (Very short, 10 words), TL = 1 (Easy, A1/A2), TC = 1 -> CVR = 1
    # Session 2: LC = 1.5 (Short, 15 words), TL = 1 (Easy, A1/A2), TC = 2 -> CVR = 3
    # Session 3: LC = 2 (Medium, 20 words), TL = 1 (Easy, A1/A2), TC = 2.5 -> CVR = 5
    # Session 4: LC = 2 (Medium, 20 words), TL = 1.5 (Medium, B1/B2), TC = 2.333 -> CVR = 7
    # Session 5: LC = 2 (Medium, 20 words), TL = 1.5 (Medium, B1/B2), TC = 3 -> CVR = 9
    # Session 6: LC = 2 (Medium, 20 words), TL = 1.5 (Medium, B1/B2), TC = 3.667 -> CVR = 11
    # Session 7: LC = 2 (Medium, 20 words), TL = 1.5 (Medium, B1/B2), TC = 4.333 -> CVR = 13
    
    session_data = {
        1: {
            "lc": 1, "tl": 1, "tc": 1, "cvr": 1, "cci_id": "cci-001",
            "items": [
                {"term_vi": "Xin chào", "term_en": "Hello", "sent_vi": "Xin chào bạn, tôi là học sinh mới của lớp.", "sent_en": "Hello my friend, I am a new student here."},
                {"term_vi": "Học sinh", "term_en": "Student", "sent_vi": "Các bạn học sinh đang vui vẻ học bài mới.", "sent_en": "The students are studying hard at school."},
                {"term_vi": "Gia đình", "term_en": "Family", "sent_vi": "Tôi rất yêu thương mọi người trong gia đình mình.", "sent_en": "I love everyone in my family very much."},
                {"term_vi": "Trường học", "term_en": "School", "sent_vi": "Ngôi trường học của tôi rất lớn và rất đẹp.", "sent_en": "My school is very big and beautiful."},
                {"term_vi": "Bạn bè", "term_en": "Friend", "sent_vi": "Chúng tôi luôn là những người bạn thân thiết nhất.", "sent_en": "We are the closest friends."},
                {"term_vi": "Ngôi nhà", "term_en": "House", "sent_vi": "Ngôi nhà nhỏ của tôi rất đẹp và ấm áp.", "sent_en": "My small house is very beautiful and cozy."},
                {"term_vi": "Quyển sách", "term_en": "Book", "sent_vi": "Tôi thích đọc quyển sách này vào mỗi buổi tối.", "sent_en": "I usually read this book every evening."}
            ]
        },
        2: {
            "lc": 1.5, "tl": 1, "tc": 2, "cvr": 3, "cci_id": "cci-002",
            "items": [
                {"term_vi": "Cảnh sát giao thông", "term_en": "Traffic police", "sent_vi": "Người cảnh sát giao thông đang làm việc chăm chỉ ở ngã tư đường phố.", "sent_en": "The traffic policeman is working hard at the street intersection."},
                {"term_vi": "Du lịch nước ngoài", "term_en": "Travel abroad / Travel overseas", "sent_vi": "Gia đình tôi có kế hoạch đi du lịch nước ngoài vào mùa hè này.", "sent_en": "My family has a plan to travel abroad this summer."},
                {"term_vi": "Cửa hàng tiện lợi", "term_en": "Convenient store", "sent_vi": "Tôi thường mua đồ ăn nhẹ tại cửa hàng tiện lợi ở gần nhà tôi.", "sent_en": "I usually buy snacks at the convenience store near my house."},
                {"term_vi": "Gần đây (địa lý)", "term_en": "Nearby / Around here / Locally", "sent_vi": "Có một quán cà phê rất đẹp mới được mở ở khu vực gần đây.", "sent_en": "There is a very nice coffee shop newly opened around here."},
                {"term_vi": "Cà vạt", "term_en": "Tie", "sent_vi": "Anh ấy thường đeo một chiếc cà vạt màu xanh khi đi làm hàng ngày.", "sent_en": "He usually wears a blue tie when going to work every day."},
                {"term_vi": "Đồ móc khóa", "term_en": "Keychain", "sent_vi": "Tôi vừa mua một cái đồ móc khóa rất dễ thương ở cửa hàng đó.", "sent_en": "I have just bought a very cute keychain at that shop."},
                {"term_vi": "Tạt vào / ra một chút", "term_en": "Pop in / out", "sent_vi": "Tôi cần phải tạt vào tiệm bánh mì một chút để mua đồ ăn sáng.", "sent_en": "I need to pop in the bakery for a moment to buy breakfast."}
            ]
        },
        3: {
            "lc": 2, "tl": 1, "tc": 2.5, "cvr": 5, "cci_id": "cci-003",
            "items": [
                {"term_vi": "Cái máy lạnh", "term_en": "AC / Air Conditioner", "sent_vi": "Cái máy lạnh mới mua ở trong phòng khách nhà tôi hoạt động rất êm ái và tiết kiệm điện.", "sent_en": "The new air conditioner in my living room runs very quietly and saves electricity."},
                {"term_vi": "Cứ thoải mái", "term_en": "Freely / feel free", "sent_vi": "Bạn cứ thoải mái ngồi nghỉ ngơi ở ghế sofa trong khi tôi đi pha một bình trà nóng nhé.", "sent_en": "Please make yourself at home and relax on the sofa while I make a pot of hot tea."},
                {"term_vi": "Kỹ sư trưởng", "term_en": "Chief engineer", "sent_vi": "Người kỹ sư trưởng tài năng đó luôn đi kiểm tra tất cả các thiết bị máy móc vào mỗi sáng sớm.", "sent_en": "That talented chief engineer always checks all the machinery equipment early every morning."},
                {"term_vi": "Kỹ năng và khả năng", "term_en": "Skill and ability", "sent_vi": "Chúng ta cần phải học tập để phát triển thêm nhiều kỹ năng và khả năng của bản thân mình.", "sent_en": "We need to study to develop more skills and abilities for ourselves."},
                {"term_vi": "Thực tập sinh", "term_en": "Intern / trainee / probationer / apprentice", "sent_vi": "Người thực tập sinh mới của công ty rất chăm chỉ và luôn cố gắng hoàn thành mọi việc tốt.", "sent_en": "The new intern at our company is very hardworking and always tries to complete all tasks well."},
                {"term_vi": "Có lẽ", "term_en": "Perhaps / maybe / most likely", "sent_vi": "Có lẽ ngày mai trời sẽ mưa to nên bạn nhớ mang theo ô khi đi ra ngoài đường nhé.", "sent_en": "Perhaps it will rain heavily tomorrow, so remember to bring an umbrella when you go out."},
                {"term_vi": "Duyệt", "term_en": "Approve / give me a go-ahead", "sent_vi": "Ban giám đốc đã chính thức phê duyệt dự án phát triển mới của chúng tôi vào chiều hôm qua.", "sent_en": "The board of directors officially approved our new development project yesterday afternoon."}
            ]
        },
        4: {
            "lc": 2, "tl": 1.5, "tc": 2.333, "cvr": 7, "cci_id": "cci-004",
            "items": [
                {"term_vi": "Dễ (kiếm tiền)", "term_en": "Easy money", "sent_vi": "Nhiều người nghĩ rằng việc bán hàng trực tuyến là cách dễ kiếm tiền nhưng thực tế rất khó khăn.", "sent_en": "Many people think that selling online is an easy money way, but the reality is very difficult."},
                {"term_vi": "Hợp đồng", "term_en": "Contract", "sent_vi": "Hai bên đối tác đã thỏa thuận xong các điều khoản và sẵn sàng chuẩn bị ký kết hợp đồng.", "sent_en": "The two partners agreed on the terms and are ready to sign the new contract."},
                {"term_vi": "Ký một cái hợp đồng", "term_en": "Sign a contract", "sent_vi": "Tôi rất vui mừng khi cuối cùng cũng được ký một cái hợp đồng lao động với đối tác lớn.", "sent_en": "I am very happy to finally sign an employment contract with the major partner."},
                {"term_vi": "Trước tiên", "term_en": "First / firstly / first off", "sent_vi": "Trước tiên bạn cần hoàn thành bản báo cáo này trước khi tham gia cuộc họp quan trọng sắp tới.", "sent_en": "First off, you need to complete this report before participating in the upcoming important meeting."},
                {"term_vi": "Thương lượng", "term_en": "Negotiate", "sent_vi": "Chúng tôi đang cố gắng thương lượng với nhà cung cấp để có được mức giá ưu đãi tốt nhất.", "sent_en": "We are trying to negotiate with the supplier to get the best discounted price."},
                {"term_vi": "Ồn ào / to mồm", "term_en": "Noisy / loud-mouthed", "sent_vi": "Những người khách ở bàn bên cạnh nói chuyện quá ồn ào làm ảnh hưởng đến mọi người xung quanh.", "sent_en": "The guests at the next table were talking too loudly, affecting everyone around them."},
                {"term_vi": "Đồng nghiệp", "term_en": "Coworker / colleague / office buddy", "sent_vi": "Tôi luôn nhận được sự hỗ trợ nhiệt tình từ các đồng nghiệp thân thiện trong văn phòng làm việc.", "sent_en": "I always receive enthusiastic support from friendly colleagues in the office."}
            ]
        },
        5: {
            "lc": 2, "tl": 1.5, "tc": 3, "cvr": 9, "cci_id": "cci-005",
            "items": [
                {"term_vi": "Quá nhanh", "term_en": "(Way) too fast", "sent_vi": "Thời gian trôi qua quá nhanh và chúng tôi cần phải tập trung để hoàn thành kế hoạch đúng hạn.", "sent_en": "Time passes too fast, and we need to focus to complete the plan on time."},
                {"term_vi": "Nhiều khi / có mấy lúc", "term_en": "Sometimes / once in a while / at times", "sent_vi": "Nhiều khi tôi cảm thấy mệt mỏi vì áp lực công việc nhưng tôi vẫn không muốn bỏ cuộc đâu.", "sent_en": "Sometimes I feel tired because of work pressure, but I still do not want to give up."},
                {"term_vi": "Tôi e là không", "term_en": "I'm afraid not", "sent_vi": "Tôi e là không thể tham gia buổi tiệc tối nay cùng các bạn vì bận việc gia đình riêng.", "sent_en": "I am afraid not to join the party tonight with you because of personal family business."},
                {"term_vi": "(Ý anh) là sao?", "term_en": "What'd you mean?", "sent_vi": "Ý anh là sao khi nói rằng chúng tôi cần phải thay đổi toàn bộ nội dung thiết kế này?", "sent_en": "What do you mean by saying that we need to change the entire content of this design?"},
                {"term_vi": "Phức tạp", "term_en": "Complicated / complex / like a mess", "sent_vi": "Quy trình giải quyết các thủ tục hành chính này thực sự rất phức tạp và mất nhiều thời gian.", "sent_en": "The process of resolving these administrative procedures is indeed very complicated and takes a lot of time."},
                {"term_vi": "Mỗi 5 năm", "term_en": "Every 5 years", "sent_vi": "Nhà máy cần tiến hành bảo trì hệ thống thiết bị quy mô lớn này mỗi năm năm một lần.", "sent_en": "The factory needs to maintain this large-scale equipment system once every five years."},
                {"term_vi": "Thông thường", "term_en": "Usually / often / always", "sent_vi": "Thông thường tôi sẽ đi bộ trong công viên gần nhà để rèn luyện sức khỏe sau giờ làm việc.", "sent_en": "Usually, I will walk in the park near my house to exercise after working hours."}
            ]
        },
        6: {
            "lc": 2, "tl": 1.5, "tc": 3.667, "cvr": 11, "cci_id": "cci-006",
            "items": [
                {"term_vi": "Chuỗi cung ứng", "term_en": "Supply chain", "sent_vi": "Chuỗi cung ứng toàn cầu đã gặp rất nhiều khó khăn lớn do ảnh hưởng nghiêm trọng của dịch bệnh.", "sent_en": "The global supply chain has faced many major difficulties due to the severe impact of the pandemic."},
                {"term_vi": "Thiếu hụt", "term_en": "Shortage / scarcity", "sent_vi": "Sự thiếu hụt nguồn nhân lực trình độ cao đang là vấn đề thách thức lớn cho các doanh nghiệp.", "sent_en": "The shortage of highly qualified human resources is currently a major challenging problem for businesses."},
                {"term_vi": "Đại dịch", "term_en": "Pandemic / epidemic", "sent_vi": "Nhiều doanh nghiệp nhỏ đã phải đóng cửa vĩnh viễn vì không thể vượt qua giai đoạn đại dịch này.", "sent_en": "Many small businesses had to shut down permanently because they could not overcome this pandemic period."},
                {"term_vi": "Chính phủ Trung Quốc", "term_en": "Chinese government", "sent_vi": "Chính phủ Trung Quốc đã ban hành một số chính sách mới nhằm kiểm soát hoạt động xuất nhập khẩu.", "sent_en": "The Chinese government has issued several new policies to strictly control import and export activities."},
                {"term_vi": "đóng cửa", "term_en": "Shut down", "sent_vi": "Cửa hàng bán lẻ đó đã quyết định đóng cửa để tập trung kinh doanh trên các nền tảng số.", "sent_en": "That retail store decided to shut down to focus on doing business on digital platforms."},
                {"term_vi": "Nỗi đau thật sự", "term_en": "Real pain point", "sent_vi": "Việc không thể tối ưu hóa chi phí sản xuất đang là nỗi đau thật sự của doanh nghiệp này.", "sent_en": "Being unable to optimize production costs is currently a real pain point for this business."},
                {"term_vi": "Thượng Hải", "term_en": "Shanghai", "sent_vi": "Thành phố Thượng Hải là trung tâm tài chính và thương mại vô cùng lớn và phát triển sầm uất.", "sent_en": "Shanghai city is an extremely large and busy financial and commercial center."}
            ]
        },
        7: {
            "lc": 2, "tl": 1.5, "tc": 4.333, "cvr": 13, "cci_id": "cci-007",
            "items": [
                {"term_vi": "Hợp cái job này", "term_en": "Fit (with) this job", "sent_vi": "Anh ấy tin rằng bản thân hoàn toàn hợp cái job này nhờ vào những kinh nghiệm tích lũy trước.", "sent_en": "He believes that he fits this job perfectly thanks to his previously accumulated experience."},
                {"term_vi": "Một cách lưu loát", "term_en": "Fluently / smoothly", "sent_vi": "Cô ấy có khả năng nói tiếng Anh một cách lưu loát sau nhiều năm học tập ở nước ngoài.", "sent_en": "She is able to speak English fluently after many years of studying abroad."},
                {"term_vi": "Một cách mạch lạc", "term_en": "Coherently", "sent_vi": "Bạn cần phải trình bày ý kiến của mình một cách mạch lạc để thuyết phục được hội đồng đánh giá.", "sent_en": "You need to present your opinions coherently to convince the evaluation board."},
                {"term_vi": "Những cổ đông", "term_en": "Stakeholders / shareholders", "sent_vi": "Ban điều hành cần phải giải trình kế hoạch phát triển mới cho những cổ đông công ty được rõ.", "sent_en": "The executive board needs to explain the new development plan for the company's stakeholders to understand."},
                {"term_vi": "Tổng đài Viettel", "term_en": "Viettel call center / switchboard", "sent_vi": "Khách hàng có thể liên hệ trực tiếp với tổng đài Viettel để được hỗ trợ xử lý sự cố.", "sent_en": "Customers can contact the Viettel call center directly for support in resolving issues."},
                {"term_vi": "Dài dòng", "term_en": "Lengthy / wordy", "sent_vi": "Bản báo cáo tài chính này quá dài dòng và không tập trung vào các điểm mấu chốt quan trọng.", "sent_en": "This financial report is too lengthy and does not focus on key important points."},
                {"term_vi": "Trung Đông", "term_en": "Middle East", "sent_vi": "Tình hình kinh tế tại khu vực Trung Đông đang có những biến động rất lớn thời gian gần đây.", "sent_en": "The economic situation in the Middle East has had very large changes recently."}
            ]
        }
    }
    
    new_item_rows = []
    
    for s_idx, s_data in session_data.items():
        lc_val = s_data["lc"]
        tl_val = s_data["tl"]
        tc_val = s_data["tc"]
        cvr_val = s_data["cvr"]
        cci_id = s_data["cci_id"]
        
        for idx, item in enumerate(s_data["items"]):
            new_item_rows.append([
                'Day 2' if s_idx == 1 else 'Day 2',  # Material (or day)
                f'Session {s_idx}',                  # Session No. (str)
                f'Number {idx + 1}',                  # Item_id
                cci_id,                              # CCI-id
                cvr_val,                             # CVR-id
                item['term_vi'],                     # Term (Tiếng Việt)
                item['term_en'],                     # Term (Tiếng Anh)
                s_idx,                               # Session No. (int)
                item['sent_vi'],                     # Complete Sentence (Vie)
                item['sent_en'],                     # Complete Sentence (Eng)
                tc_val,                              # TC
                lc_val,                              # LC
                tl_val                               # TL
            ])
            
    # Create the new workbook
    wb_dest = Workbook()
    
    # Sheet 1: Chunks-resource - CVR_new
    ws_items = wb_dest.active
    ws_items.title = 'Chunks-resource - CVR_new'
    ws_items.append(item_headers)
    for r in new_item_rows:
        ws_items.append(r)
        
    # Sheet 2: Package-test
    ws_pkg = wb_dest.create_sheet(title='Package-test')
    ws_pkg.append(package_headers)
    for r in package_rows:
        ws_pkg.append(r)
        
    # Sheet 3: CCI
    ws_cci = wb_dest.create_sheet(title='CCI')
    ws_cci.append(cci_headers)
    for r in cci_rows:
        ws_cci.append(r)
        
    # Save the new workbook
    wb_dest.save(dest_path)
    print(f"Successfully generated new package workbook: {dest_path}")
    print(f"Total sessions: 7")
    print(f"Total items added: {len(new_item_rows)}")

if __name__ == '__main__':
    main()
