import csv
import openpyxl
from openpyxl import Workbook

def main():
    csv_path = 'chunks-resourcce/Package-Green-test - Sessions list.csv'
    dest_path = 'chunks-resourcce/Chunks Resource.xlsx'
    
    # Natural sentences and exact parameter values:
    # Session 1: Target CVR = 1, LC = 1.0 (10 words), TL = 1.0 (Easy A1/A2), TC = 1.0. Simpler everyday words.
    # Session 2: Target CVR = 3, LC = 1.5 (15 words), TL = 1.0 (Easy A1/A2), TC = 2.0.
    # Session 3: Target CVR = 5, LC = 1.5 (15 words), TL = 1.0 (Easy A1/A2), TC = 3.333.
    # Session 4: Target CVR = 7, LC = 1.8 (18 words), TL = 1.5 (Medium B1/B2), TC = 2.593.
    # Session 5: Target CVR = 9, LC = 1.8 (18 words), TL = 1.5 (Medium B1/B2), TC = 3.0.
    # Session 6: Target CVR = 11, LC = 1.8 (18 words), TL = 1.5 (Medium B1/B2), TC = 3.667.
    # Session 7: Target CVR = 13, LC = 2.1 (21 words, < 22), TL = 1.5 (Medium B1/B2), TC = 4.127.
    
    session_data = {
        1: {
            "lc": 1.0, "tl": 1.0, "tc": 1.0, "cvr": 1, "cci_id": "cci-001",
            "items": [
                {"term_vi": "Mẹ", "term_en": "Mother", "sent_vi": "Mẹ tôi luôn nấu những món ăn ngon cho cả nhà.", "sent_en": "My mother always cooks delicious dishes for the whole family."},
                {"term_vi": "Bố", "term_en": "Father", "sent_vi": "Bố tôi thường chở tôi đi học vào mỗi buổi sáng.", "sent_en": "My father usually takes me to school every morning."},
                {"term_vi": "Mèo", "term_en": "Cat", "sent_vi": "Con mèo nhỏ của tôi thích nằm sưởi nắng ngoài sân.", "sent_en": "My small cat likes to lie in the sun in the yard."},
                {"term_vi": "Chó", "term_en": "Dog", "sent_vi": "Chú chó cưng đang chạy nhảy vui vẻ ở ngoài sân.", "sent_en": "The pet dog is running and playing happily in the yard."},
                {"term_vi": "Nhà", "term_en": "House", "sent_vi": "Ngôi nhà nhỏ của tôi luôn tràn ngập niềm vui.", "sent_en": "My small house is always filled with joy."},
                {"term_vi": "Ăn", "term_en": "Eat", "sent_vi": "Tôi thích ăn bánh mì kẹp thịt vào bữa sáng.", "sent_en": "I like to eat bread with meat for breakfast."},
                {"term_vi": "Táo", "term_en": "Apple", "sent_vi": "Tôi vừa mua mấy quả táo rất tươi ở chợ.", "sent_en": "I just bought some very fresh apples at the market."}
            ]
        },
        2: {
            "lc": 1.5, "tl": 1.0, "tc": 2.0, "cvr": 3, "cci_id": "cci-002",
            "items": [
                {"term_vi": "Duyệt", "term_en": "Approve / give me a go-ahead", "sent_vi": "Sếp đã duyệt kế hoạch cho chuyến đi du lịch của phòng chúng ta rồi.", "sent_en": "The boss has approved the plan for our department's trip already."},
                {"term_vi": "..., có thể nói vậy", "term_en": "..., so to speak", "sent_vi": "Công ty này giống như ngôi nhà thứ hai của tôi, có thể nói vậy.", "sent_en": "This company is like a second home to me, so to speak."},
                {"term_vi": "Cải thiện", "term_en": "Work on / improve", "sent_vi": "Tôi đang cố gắng luyện nói hàng ngày để cải thiện vốn tiếng Anh của mình.", "sent_en": "I am trying to practice speaking every day to improve my English."},
                {"term_vi": "Quản trị", "term_en": "Management", "sent_vi": "Học ngành quản trị kinh doanh sẽ giúp bạn có nhiều cơ hội việc làm.", "sent_en": "Studying business management will help you have many job opportunities."},
                {"term_vi": "Nguyên tắc", "term_en": "Philosophy / principle", "sent_vi": "Sống có nguyên tắc sẽ giúp bạn dễ dàng đạt được những mục tiêu của mình.", "sent_en": "Living with principles will help you easily achieve your goals."},
                {"term_vi": "Dài dòng", "term_en": "Lengthy / wordy", "sent_vi": "Đừng viết mail quá dài dòng vì người ta không có thời gian đọc đâu.", "sent_en": "Don't write emails too long because people don't have time to read them all."},
                {"term_vi": "Dầu gió", "term_en": "Medicated oil", "sent_vi": "Mỗi khi bị nhức đầu tôi lại thích xoa một ít dầu gió cho đỡ.", "sent_en": "Whenever I have a headache, I like to rub some medicated oil to feel better."}
            ]
        },
        3: {
            "lc": 1.5, "tl": 1.0, "tc": 3.333, "cvr": 5, "cci_id": "cci-003",
            "items": [
                {"term_vi": "Cái máy lạnh", "term_en": "AC / Air Conditioner", "sent_vi": "Cái máy lạnh nhà tôi bị hỏng đúng vào những ngày nắng nóng thế này.", "sent_en": "My air conditioner broke down right in these hot sunny days."},
                {"term_vi": "Cứ thoải mái", "term_en": "Freely / feel free", "sent_vi": "Đến nhà tôi chơi thì cứ thoải mái tự nhiên như ở nhà mình nhé.", "sent_en": "When visiting my house, please make yourself at home."},
                {"term_vi": "Kỹ sư trưởng", "term_en": "Chief engineer", "sent_vi": "Bố tôi hiện đang làm kỹ sư trưởng cho một tập đoàn xây dựng lớn.", "sent_en": "My father is currently working as a chief engineer for a large construction corporation."},
                {"term_vi": "Kỹ năng và khả năng", "term_en": "Skill and ability", "sent_vi": "Công việc này đòi hỏi bạn phải có cả kỹ năng và khả năng giao tiếp.", "sent_en": "This job requires you to have both communication skills and abilities."},
                {"term_vi": "Thực tập sinh", "term_en": "Intern / trainee / probationer / apprentice", "sent_vi": "Công ty chúng tôi đang tuyển thêm một số thực tập sinh cho dự án mới.", "sent_en": "Our company is recruiting some more interns for the new project."},
                {"term_vi": "Có lẽ", "term_en": "Perhaps / maybe / most likely", "sent_vi": "Có lẽ tôi nên xin nghỉ một ngày để đi khám sức khỏe xem sao.", "sent_en": "Perhaps I should take a day off to check my health."},
                {"term_vi": "Duyệt", "term_en": "Approve / give me a go-ahead", "sent_vi": "Hồ sơ của cậu vẫn đang chờ ban giám đốc duyệt nên cứ từ từ.", "sent_en": "Your profile is still waiting for the directors to approve, so take it easy."}
            ]
        },
        4: {
            "lc": 1.8, "tl": 1.5, "tc": 2.593, "cvr": 7, "cci_id": "cci-004",
            "items": [
                {"term_vi": "Dễ (kiếm tiền)", "term_en": "Easy money", "sent_vi": "Làm gì có công việc nào vừa nhàn hạ lại vừa dễ kiếm tiền như cậu nghĩ đâu.", "sent_en": "There is no job that is both easy and easy money as you think."},
                {"term_vi": "Hợp đồng", "term_en": "Contract", "sent_vi": "Cậu nhớ đọc kỹ các điều khoản trước khi ký vào bản hợp đồng kinh tế này nhé.", "sent_en": "Remember to read the terms carefully before signing this economic contract."},
                {"term_vi": "Ký một cái hợp đồng", "term_en": "Sign a contract", "sent_vi": "Cuối cùng thì chúng tôi cũng đã ký một cái hợp đồng hợp tác với đối tác này.", "sent_en": "Finally we have signed a cooperation contract with this partner."},
                {"term_vi": "Trước tiên", "term_en": "First / firstly / first off", "sent_vi": "Trước tiên hãy dọn dẹp phòng làm việc của cậu cho thật sạch sẽ và gọn gàng đã.", "sent_en": "First of all, clean up your office room so that it is neat and tidy."},
                {"term_vi": "Thương lượng", "term_en": "Negotiate", "sent_vi": "Chúng tôi đang cố gắng thương lượng để mua được lô hàng với mức giá ưu đãi nhất.", "sent_en": "We are trying to negotiate to buy the batch at the most discounted price."},
                {"term_vi": "Ồn ào / to mồm", "term_en": "Noisy / loud-mouthed", "sent_vi": "Mấy người ở bàn bên cạnh cứ nói chuyện ồn ào làm tôi không tập trung được đây.", "sent_en": "The people at the next table kept talking noisily, making me unable to focus."},
                {"term_vi": "Đồng nghiệp", "term_en": "Coworker / colleague / office buddy", "sent_vi": "Tôi rất may mắn khi được làm việc chung với những người đồng nghiệp vô cùng thân thiết.", "sent_en": "I am very lucky to work together with extremely friendly colleagues."}
            ]
        },
        5: {
            "lc": 1.8, "tl": 1.5, "tc": 3.0, "cvr": 9, "cci_id": "cci-005",
            "items": [
                {"term_vi": "Nhiều khi / có mấy lúc", "term_en": "Sometimes / once in a while / at times", "sent_vi": "Nhiều khi tôi chỉ muốn vứt hết công việc để đi du lịch một chuyến thật xa thôi.", "sent_en": "Sometimes I just want to drop all work to go on a trip far away."},
                {"term_vi": "Tôi e là không", "term_en": "I'm afraid not", "sent_vi": "Tôi e là không thể hoàn thành dự án này đúng thời hạn mà sếp yêu cầu đâu.", "sent_en": "I'm afraid not to be able to complete this project within the deadline requested by the boss."},
                {"term_vi": "(Ý anh) là sao?", "term_en": "What'd you mean?", "sent_vi": "Ý anh là sao khi bảo rằng phương án thiết kế của chúng tôi không hợp lý vậy?", "sent_en": "What do you mean by saying that our design option is not reasonable?"},
                {"term_vi": "Phức tạp", "term_en": "Complicated / complex / like a mess", "sent_vi": "Vấn đề này xem ra rất phức tạp chứ không hề dễ giải quyết như cậu nghĩ đâu.", "sent_en": "This problem seems very complicated, not at all easy to solve as you think."},
                {"term_vi": "Mỗi 5 năm", "term_en": "Every 5 years", "sent_vi": "Mỗi năm năm một lần, công ty chúng tôi lại tổ chức kỳ thi nâng bậc thợ lớn.", "sent_en": "Once every five years, our company organizes a major worker upgrading exam."},
                {"term_vi": "Thông thường", "term_en": "Usually / often / always", "sent_vi": "Thông thường tôi sẽ tự nấu ăn ở nhà chứ ít khi ra ngoài hàng quán ăn uống.", "sent_en": "Usually, I will cook at home myself, rarely going out to eat at restaurants."},
                {"term_vi": "Vấn đề thiếu hụt nhân lực", "term_en": "Staff shortage / headcount problem", "sent_vi": "Vấn đề thiếu hụt nhân lực đang khiến bộ phận sản xuất gặp rất nhiều khó khăn lớn.", "sent_en": "The staff shortage problem is causing the production department to face many difficulties."}
            ]
        },
        6: {
            "lc": 1.8, "tl": 1.5, "tc": 3.667, "cvr": 11, "cci_id": "cci-006",
            "items": [
                {"term_vi": "Tốt hơn là cậu nên…", "term_en": "You'd better… / You should...", "sent_vi": "Tốt hơn là cậu nên chủ động xin lỗi sếp trước khi sự việc tệ hơn rất nhiều.", "sent_en": "You'd better actively apologize to the boss before things get much worse."},
                {"term_vi": "Chuỗi cung ứng", "term_en": "Supply chain", "sent_vi": "Chuỗi cung ứng của tập đoàn bị gián đoạn nghiêm trọng do ảnh hưởng dịch bệnh vừa qua.", "sent_en": "The group's supply chain was severely disrupted due to the impact of the recent epidemic."},
                {"term_vi": "Thiếu hụt", "term_en": "Shortage / scarcity", "sent_vi": "Sự thiếu hụt nguồn nước sạch vào mùa khô đang là nỗi lo lớn của người dân đây.", "sent_en": "The shortage of clean water during the dry season is currently a big worry for the people."},
                {"term_vi": "Đại dịch", "term_en": "Pandemic / epidemic", "sent_vi": "Rất nhiều cửa hàng bán lẻ đã phải đóng cửa vĩnh viễn sau thời kỳ đại dịch đó.", "sent_en": "A lot of retail shops had to close permanently after that pandemic period."},
                {"term_vi": "Chính phủ Trung Quốc", "term_en": "Chinese government", "sent_vi": "Chính phủ Trung Quốc đang nỗ lực thực hiện các biện pháp giúp phục hồi kinh tế nước.", "sent_en": "The Chinese government is striving to implement measures to help recover the economy."},
                {"term_vi": "đóng cửa", "term_en": "Shut down", "sent_vi": "Nhà máy quyết định đóng cửa một số phân xưởng hoạt động không có hiệu quả kinh tế.", "sent_en": "The factory decided to close down some workshops that are not economically effective."},
                {"term_vi": "Nỗi đau thật sự", "term_en": "Real pain point", "sent_vi": "Việc không có khách hàng trung thành chính là nỗi đau thật sự của doanh nghiệp nhỏ này.", "sent_en": "Not having loyal customers is the real pain point for this small business."}
            ]
        },
        7: {
            "lc": 2.1, "tl": 1.5, "tc": 4.127, "cvr": 13, "cci_id": "cci-007",
            "items": [
                {"term_vi": "Trào ngược dạ dày", "term_en": "Acid reflux", "sent_vi": "Chứng trào ngược dạ dày thường xuyên xảy ra nếu bạn ăn quá no và nằm nghỉ ngay sau bữa ăn.", "sent_en": "Acid reflux disease often occurs if you eat too full and lie down to rest right after the meal."},
                {"term_vi": "Ợ nóng", "term_en": "Heartburn", "sent_vi": "Triệu chứng ợ nóng thường xuyên xuất hiện nếu bạn có thói quen ăn đồ ăn nhiều dầu mỡ buổi tối muộn nhé.", "sent_en": "Heartburn symptoms often appear if you have a habit of eating greasy food late at night."},
                {"term_vi": "Đường huyết", "term_en": "Blood sugar", "sent_vi": "Chúng ta cần phải chú ý theo dõi chỉ số đường huyết để duy trì sức khỏe bằng chế độ dinh dưỡng tốt.", "sent_en": "We need to pay attention to monitoring the blood sugar index to maintain good health with a good diet."},
                {"term_vi": "Ngộ độc thực phẩm", "term_en": "Food poisoning", "sent_vi": "Một vụ ngộ độc thực phẩm nghiêm trọng xảy ra tại quán ăn lớn đó đã làm nhiều người phải đi cấp cứu.", "sent_en": "A serious food poisoning incident that occurred at that large restaurant sent many people to emergency treatment."},
                {"term_vi": "Ngấn mỡ bụng", "term_en": "Love handles", "sent_vi": "Chế độ tập luyện thể thao đều đặn mỗi ngày sẽ giúp bạn đốt cháy mỡ thừa và loại bỏ ngấn mỡ bụng.", "sent_en": "Regular daily sports exercise will help you burn fat and remove love handles quickly."},
                {"term_vi": "\"Có tâm\"", "term_en": "Dedicated", "sent_vi": "Tôi rất quý mến người sếp rất có tâm vì anh ấy luôn chỉ bảo công việc cho nhân viên mới.", "sent_en": "I really like that very dedicated boss because he always guides the new employees in their work."},
                {"term_vi": "Còn khuya mới...", "term_en": "Nowhere near... / far from...", "sent_vi": "Học sinh lười biếng kia còn khuya mới thi đỗ nếu không chịu thay đổi thái độ học tập từ nay.", "sent_en": "That lazy student is nowhere near passing the exam if he doesn't change his study attitude from now."}
            ]
        }
    }
    
    # 2. Build CCI definitions
    cci_headers = ['Session', 'CCI_id', 'CCI Name', 'Ampe (A)', 'Description', 'Category']
    cci_rows = [
        [1, 'cci-001', 'Give it a shot', 2, 'Linear 1 on 1 as Blow', 'Blow'],
        [2, 'cci-002', 'Go with the flow', 2, 'Linear RPD-free as Flow', 'Flow'],
        [3, 'cci-003', 'Chunks on the go', 4, 'Linear chunking act as Chunks', 'Chunks'],
        [4, 'cci-004', 'Freeze', 4, 'Freeze your body with RPD-free or 1-on-1 sound', 'null'],
        [5, 'cci-005', 'Robot', 6, 'Move your hands linearly 1-on-1 as Blow', 'Blow'],
        [6, 'cci-006', 'Taichi', 6, 'Move your hands nonstop freely as Flow', 'Flow'],
        [7, 'cci-007', 'Strike', 8, 'Strike a fixed n times as Chunks', 'Chunks']
    ]
    
    # 3. Build Package-test definitions
    package_headers = ['Package_id', 'Name', 'Description', 'Session list', 'CCI list']
    package_rows = []
    for i in range(1, 8):
        package_rows.append([
            'Package-Green-test',
            f'Test {i:02d}',
            'Bộ test Green 7 sessions',
            f'session_id -{i}',
            f'cci-id {i:02d}'
        ])
        
    # 4. Build Chunks-resource items
    item_headers = [
        'Material',
        'Session No.',
        'Item_id',
        'CCI-id',
        'CVR-id',
        'Term (Tiếng Việt)',
        'Term (Tiếng Anh)',
        'Session No.',
        'Complete Sentence (Vie)',
        'Complete Sentence (Eng)',
        'TC',
        'LC',
        'TL'
    ]
    
    new_item_rows = []
    for s_idx in range(1, 8):
        s_data = session_data[s_idx]
        lc_val = s_data["lc"]
        tl_val = s_data["tl"]
        tc_val = s_data["tc"]
        cvr_val = s_data["cvr"]
        cci_id = s_data["cci_id"]
        
        for idx, item in enumerate(s_data["items"]):
            new_item_rows.append([
                'Day 2',                  # Material
                f'Session {s_idx}',        # Session No. (str)
                f'Number {idx + 1}',        # Item_id
                cci_id,                    # CCI-id
                cvr_val,                   # CVR-id
                item['term_vi'],           # Term (Tiếng Việt)
                item['term_en'],           # Term (Tiếng Anh)
                s_idx,                     # Session No. (int)
                item['sent_vi'],           # Complete Sentence (Vie)
                item['sent_en'],           # Complete Sentence (Eng)
                tc_val,                    # TC
                lc_val,                    # LC
                tl_val                     # TL
            ])
            
    # Create the new workbook
    wb_dest = Workbook()
    
    # Sheet 1: Chunks-resource - CVR_new
    ws_items = wb_dest.active
    ws_items.title = 'Chunks-resource - CVR_new'
    ws_items.append(item_headers)
    for r in new_item_rows:
        ws_items.append(r)
        
    # Sheet 2: Package-test
    ws_pkg = wb_dest.create_sheet(title='Package-test')
    ws_pkg.append(package_headers)
    for r in package_rows:
        ws_pkg.append(r)
        
    # Sheet 3: CCI
    ws_cci = wb_dest.create_sheet(title='CCI')
    ws_cci.append(cci_headers)
    for r in cci_rows:
        ws_cci.append(r)
        
    # Save the workbook to Chunks Resource.xlsx
    wb_dest.save(dest_path)
    print(f"Successfully generated final Green workbook at: {dest_path}")
    print(f"Total sessions: 7")
    print(f"Total items added: {len(new_item_rows)}")

if __name__ == '__main__':
    main()
